from __future__ import annotations

from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple
import os
import re
import requests
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

TEMPLATE_PATH = Path(__file__).with_name("burn_plan_template.xlsx")

CELL_MAP = {
    "tract_name": "B2", "burn_address": "B3", "state": "B4", "county": "E4",
    "burn_mgr_name": "B5", "burn_mgr_cert": "G5", "burn_mgr_phone": "B6",
    "executers_mailing_address": "B7", "prepared_by": "B8", "date_prepared": "G9",
    "burn_county": "B12", "lat_long": "B13", "burn_acres": "B14",
    "overstory_type": "B15", "understory_type": "B16", "fuel_type_amount": "B17",
    "topography": "B18", "special_features": "A20", "objectives": "A22",
    "manpower_equipment": "A25", "adversely_affected_areas": "A27",
    "breach_potential": "A29", "smoke_precautions": "A31", "emergency_resources": "A33",
    "desired_surface_wind": "E35", "forecast_surface_wind": "G35", "observed_surface_wind": "I35",
    "desired_humidity": "E36", "forecast_humidity": "G36", "observed_humidity": "I36",
    "desired_temperature": "E37", "forecast_temperature": "G37", "observed_temperature": "I37",
    "desired_transport_wind": "E38", "forecast_transport_wind": "G38", "observed_transport_wind": "I38",
    "desired_mixing_height": "E39", "forecast_mixing_height": "G39", "observed_mixing_height": "I39",
    "desired_dispersion_index": "E40", "forecast_dispersion_index": "G40", "observed_dispersion_index": "I40",
    "desired_fine_fuel_moisture": "E41", "forecast_fine_fuel_moisture": "G41", "observed_fine_fuel_moisture": "I41",
    "desired_kbdi": "E42", "forecast_kbdi": "G42", "observed_kbdi": "I42",
    "start_time": "B44", "completion_time": "E44", "hours_to_complete": "B45", "flame_length": "F45",
    "ignition_techniques": "B46", "permit_number": "B48", "actual_burn_date": "G48",
}

@dataclass
class BurnInputs:
    tract_name: str = ""
    burn_address: str = ""
    state: str = "AL"
    county: str = ""
    burn_mgr_name: str = ""
    burn_mgr_cert: str = ""
    burn_mgr_phone: str = ""
    executers_mailing_address: str = ""
    prepared_by: str = ""
    latitude: float | None = None
    longitude: float | None = None
    burn_acres: float | None = None
    burn_type: str = ""
    overstory_type: str = ""
    understory_type: str = ""
    fuel_type_amount: str = ""
    topography: str = ""
    special_features: str = ""
    objectives: str = ""
    smoke_sensitive_areas: str = ""
    water_sources: str = ""
    roads_access: str = ""
    neighbors: str = ""
    manpower_equipment: str = ""
    nighttime_smoke_screening: str = ""
    special_precautions: str = ""
    breach_potential: str = ""
    smoke_precautions: str = ""
    emergency_resources: str = ""
    ignition_techniques: str = ""
    desired_surface_wind: str = ""
    desired_humidity: str = ""
    desired_temperature: str = ""
    desired_transport_wind: str = ""
    desired_mixing_height: str = ""
    desired_dispersion_index: str = ""
    desired_fine_fuel_moisture: str = ""
    desired_kbdi: str = ""
    observed_surface_wind: str = ""
    observed_humidity: str = ""
    observed_temperature: str = ""
    observed_transport_wind: str = ""
    observed_mixing_height: str = ""
    observed_dispersion_index: str = ""
    observed_fine_fuel_moisture: str = ""
    observed_kbdi: str = ""
    hours_to_complete: str = ""
    flame_length: str = ""
    start_time: str = ""
    completion_time: str = ""
    permit_number: str = ""
    actual_burn_date: str = ""
    prepared_by_name: str = ""
    prepared_by_date: str = ""
    witnessed_by_name: str = ""
    witnessed_by_date: str = ""

@dataclass
class WeatherInputs:
    # NWS county forecast values. These never replace the desired prescription
    # or observed day-of-burn conditions.
    surface_wind_mph: float | None = None
    surface_wind_dir: str = ""
    min_rh: float | None = None
    max_temp_f: float | None = None
    transport_wind_mph: float | None = None
    transport_wind_dir: str = ""
    mixing_height_ft: float | None = None
    dispersion_index: float | None = None
    kbdi: float | None = None
    forecast_period: str = ""
    forecast_county: str = ""
    forecast_office: str = ""
    forecast_issued: str = ""
    forecast_product_id: str = ""
    chance_precip_pct: float | None = None
    precip_type: str = ""
    precip_amount: str = ""
    stability_class: str = ""
    max_lvori: float | None = None
    dispersion_category: str = ""
    remarks: str = ""

PRESCRIPTION_TEMPLATES: Dict[str, Dict[str, str]] = {
    "Rangeland": {
        "season": "Restoration: Jan-Dec; Maintenance: Dec-March",
        "desired_temperature": "Restoration 45-95°F; Maintenance 45-70°F",
        "desired_humidity": "Restoration 30-60%; Maintenance 30-50%",
        "desired_surface_wind": "2-6 mph midflame",
        "desired_transport_wind": "9 mph +",
        "desired_mixing_height": "1650 ft +",
        "desired_dispersion_index": "41-60",
        "desired_fine_fuel_moisture": "1-hour 6-12%; 10-hour 10-15% or higher",
        "desired_kbdi": "Restoration <450; Maintenance <300",
        "ignition_techniques": "Flank, strip-head, head, ring, grid, or ridge; drip torch, powertorch, drone, helicopter, or pyroshot.",
        "flame_length": "Restoration 2-8 ft; Maintenance 3-12 in",
        "nighttime_viable": "No",
        "notes": "Watch for smoldering if mastication was completed first and thatch under green grass.",
    },
    "Fuel Reduction": {
        "season": "Generally dormant season; Nov-April common",
        "desired_temperature": "50-70°F",
        "desired_humidity": "30-55%",
        "desired_surface_wind": "2-6 mph midflame",
        "desired_transport_wind": "9 mph +",
        "desired_mixing_height": "1650 ft +",
        "desired_dispersion_index": "41-60",
        "desired_fine_fuel_moisture": "1-hour 6-10%; 10-hour 10-15%",
        "desired_kbdi": "<300",
        "ignition_techniques": "Flanking, grid, strip-head, or head fire with higher 1-hour fuel moisture/RH; drip torch, powertorch, drone, helicopter, or pyroshot.",
        "flame_length": "2-4 ft typical for sapling fuel reduction; keep lower where timber scorch risk exists.",
        "nighttime_viable": "Yes",
        "notes": "Review whether the unit has been burned in four or more years and manage heavy fuel pockets conservatively.",
    },
    "Pre-Marking": {
        "season": "Nov-April",
        "desired_temperature": "45-75°F",
        "desired_humidity": "30-55%",
        "desired_surface_wind": "<2 mph in overstocked stands",
        "desired_transport_wind": "9 mph +",
        "desired_mixing_height": "1650 ft +",
        "desired_dispersion_index": "41-60",
        "desired_fine_fuel_moisture": "1-hour 6-12%; 10-hour 15% +",
        "desired_kbdi": "<300",
        "ignition_techniques": "Backing, flanking, grid, strip-head, or ridge ignition; drip torch, powertorch, drone, pyroshot, or helicopter.",
        "flame_length": "1-3 ft",
        "nighttime_viable": "Yes",
        "notes": "Confirm landowner tolerance to scorch before ignition.",
    },
    "TSI": {
        "season": "Depends on stand: thinned pine May-Oct; longleaf Jan-Dec; young longleaf Dec-early March; upland mixed hardwood Dec-March",
        "desired_temperature": "45-100°F depending on stand type",
        "desired_humidity": "30-65% depending on stand type",
        "desired_surface_wind": "Use conservative midflame winds suited to stand/fuels",
        "desired_transport_wind": "9 mph +",
        "desired_mixing_height": "1650 ft +",
        "desired_dispersion_index": "41-60",
        "desired_fine_fuel_moisture": "1-hour 6-14%; 10-hour 12-15% + depending on stand",
        "desired_kbdi": "<300 to <450 depending on stand and season",
        "ignition_techniques": "Backing, flanking, grid, strip-head, head, or ridge ignition; drip torch, powertorch, drone, pyroshot, or helicopter.",
        "flame_length": "<3 ft for thinned pine; 2-4 ft for longleaf; 4-12 in for young longleaf; <2 ft for upland mixed hardwood.",
        "nighttime_viable": "Depends on objective and stand; yes for some TSI, no for young longleaf/upland hardwood recommendations.",
        "notes": "Stand-specific review is required. Consider brush density, longleaf bud/candle stage, and hardwood timber quality/scorch tolerance.",
    },
    "Site Prep": {
        "season": "Objective and site dependent",
        "desired_temperature": "Site/objective dependent",
        "desired_humidity": "Use prescription approved by burn manager",
        "desired_surface_wind": "Use prescription approved by burn manager",
        "desired_transport_wind": "9 mph + preferred",
        "desired_mixing_height": "1650 ft +",
        "desired_dispersion_index": "41-60 preferred",
        "desired_fine_fuel_moisture": "Site/objective dependent",
        "desired_kbdi": "Site/objective dependent",
        "ignition_techniques": "Select firing method based on slash, desired intensity, smoke, and holding concerns.",
        "flame_length": "Site/objective dependent",
        "nighttime_viable": "Manager review required",
        "notes": "Site prep prescription should be adjusted for slash loading, soil/site sensitivity, crop-tree concerns, and smoke receptors.",
    },
    "Wildlife": {
        "season": "Objective dependent; dormant or growing season may be appropriate",
        "desired_temperature": "Objective dependent",
        "desired_humidity": "30-60% typical planning range",
        "desired_surface_wind": "2-6 mph midflame commonly used for low-to-moderate intensity burns",
        "desired_transport_wind": "9 mph + preferred",
        "desired_mixing_height": "1650 ft +",
        "desired_dispersion_index": "41-60 preferred",
        "desired_fine_fuel_moisture": "Objective dependent",
        "desired_kbdi": "Objective dependent",
        "ignition_techniques": "Backing, flanking, strip-head, grid, or ring fire depending on desired patchiness and intensity.",
        "flame_length": "Objective dependent; keep within acceptable scorch and escape-risk limits.",
        "nighttime_viable": "Manager review required",
        "notes": "Match season and intensity to target wildlife habitat response, nesting considerations, and fuel conditions.",
    },
}

def get_prescription_template(burn_type: str) -> Dict[str, str]:
    return PRESCRIPTION_TEMPLATES.get(burn_type, {})

def desired_conditions(burn_type: str = "") -> Dict[str, str]:
    t = get_prescription_template(burn_type)
    defaults = {
        "desired_surface_wind": "5-15 mph, steady direction",
        "desired_humidity": "30-55% typical; avoid extreme low RH",
        "desired_temperature": "Site/objective dependent",
        "desired_transport_wind": "9-20 mph preferred",
        "desired_mixing_height": "> 1700 ft",
        "desired_dispersion_index": "> 26; use caution > 100",
        "desired_fine_fuel_moisture": "Approx. RH / 5",
        "desired_kbdi": "<300 dormant; <450 growing; <550 site prep",
    }
    defaults.update({k: v for k, v in t.items() if k in defaults})
    return defaults

def build_rule_check(weather: WeatherInputs) -> List[Tuple[str, str, str]]:
    checks: List[Tuple[str, str, str]] = []
    def ok(flag: bool, item: str, good: str, bad: str):
        checks.append(("OK" if flag else "REVIEW", item, good if flag else bad))
    if weather.surface_wind_mph is not None:
        ok(5 <= weather.surface_wind_mph <= 15, "Surface wind", "Within 5-15 mph preferred range.", "Outside 5-15 mph preferred range.")
    if weather.min_rh is not None:
        ok(25 <= weather.min_rh <= 60, "Relative humidity", "Within broad planning range.", "RH may be too low/high; review objective and fuels.")
    if weather.transport_wind_mph is not None:
        ok(weather.transport_wind_mph >= 9, "Transport wind", "At least 9 mph.", "Below 9 mph recommendation in prescription references.")
    if weather.mixing_height_ft is not None:
        ok(weather.mixing_height_ft >= 1650, "Mixing height", "At least 1,650 ft.", "Below 1,650 ft recommendation in prescription references.")
    if weather.dispersion_index is not None:
        ok(41 <= weather.dispersion_index <= 60, "Dispersion index", "Within 41-60 target range.", "Outside 41-60 target range; review smoke management.")
    if weather.kbdi is not None:
        ok(weather.kbdi < 450, "KBDI", "Within conservative planning range.", "High KBDI; review season/objective and mop-up needs.")
    if not checks:
        checks.append(("INFO", "Weather", "Enter forecast/observed weather to run checks."))
    return checks

def basic_ai_draft(inputs: BurnInputs, weather: WeatherInputs) -> Dict[str, str]:
    t = get_prescription_template(inputs.burn_type)
    objective = inputs.objectives or f"Conduct a prescribed burn for {inputs.burn_type or 'the selected management objective'} while maintaining control, smoke management, and resource protection."
    if t.get("notes"):
        objective += f" Prescription note: {t['notes']}"
    return {
        "special_features": inputs.special_features or "Protect marked utilities, structures, boundary lines, SMZs, gates, cultural resources, wildlife structures, and other sensitive resources identified during pre-burn inspection.",
        "objectives": objective,
        "manpower_equipment": inputs.manpower_equipment or "Crew and equipment shall be sized to acreage, fuel loading, weather, holding needs, and contingency requirements. Minimum resources should include burn boss, ignition and holding personnel, water, communications, PPE, hand tools, and mop-up resources.",
        "adversely_affected_areas": f"Nighttime smoke screening: {inputs.nighttime_smoke_screening or 'Not specified'}. Special precautions: {inputs.special_precautions or 'None listed'}.",
        "breach_potential": inputs.breach_potential or "Review downwind lines, corners, heavy fuel pockets, road edges, and changes in topography. Strengthen weak line sections before ignition and assign holding resources to high-risk points.",
        "smoke_precautions": inputs.smoke_precautions or "Burn only with favorable transport/surface winds, adequate mixing height, and acceptable dispersion. Monitor smoke-sensitive receptors and stop ignition if smoke impacts become unsafe.",
        "emergency_resources": inputs.emergency_resources or "Confirm permit requirements, 911 access, local fire department, nearest hospital, law enforcement/traffic needs, water sources, and evacuation/access routes before ignition.",
        "ignition_techniques": inputs.ignition_techniques or t.get("ignition_techniques", "Use backing fire along control lines, then flanking/strip-head fire as conditions allow. Adjust firing pattern to maintain desired flame length, smoke lift, and holding safety."),
    }

def optional_openai_polish(draft: Dict[str, str], inputs: BurnInputs, weather: WeatherInputs) -> Dict[str, str]:
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        return draft
    try:
        from openai import OpenAI
        import json
        client = OpenAI(api_key=api_key)
        resp = client.responses.create(
            model="gpt-5.4-mini",
            input=[{"role": "user", "content": "Polish these burn plan fields for professional forester review. Keep conservative. Return JSON same keys only. " + str({"inputs": asdict(inputs), "weather": asdict(weather), "draft": draft})}],
            text={"format": {"type": "json_object"}},
        )
        polished = json.loads(resp.output_text)
        return {k: str(polished.get(k, v)) for k, v in draft.items()}
    except Exception:
        return draft

def fill_template(inputs: BurnInputs, weather: WeatherInputs, output_path: str | Path, use_ai: bool = False) -> Path:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb.active
    data = asdict(inputs)
    data["burn_county"] = inputs.county
    data["lat_long"] = "" if inputs.latitude is None or inputs.longitude is None else f"{inputs.latitude:.6f}, {inputs.longitude:.6f}"
    data["date_prepared"] = datetime.now().strftime("%m/%d/%Y")
    draft = basic_ai_draft(inputs, weather)
    if use_ai:
        draft = optional_openai_polish(draft, inputs, weather)
    data.update(draft)
    for k, v in desired_conditions(inputs.burn_type).items():
        if not data.get(k): data[k] = v
    if weather.surface_wind_mph is not None: data["forecast_surface_wind"] = f"{weather.surface_wind_mph:g} mph {weather.surface_wind_dir}".strip()
    if weather.min_rh is not None:
        data["forecast_humidity"] = f"{weather.min_rh:g}%"; data["forecast_fine_fuel_moisture"] = f"~{weather.min_rh/5:.1f}%"
    if weather.max_temp_f is not None: data["forecast_temperature"] = f"{weather.max_temp_f:g}°F"
    if weather.transport_wind_mph is not None: data["forecast_transport_wind"] = f"{weather.transport_wind_mph:g} mph {weather.transport_wind_dir}".strip()
    if weather.mixing_height_ft is not None: data["forecast_mixing_height"] = f"{weather.mixing_height_ft:g} ft"
    if weather.dispersion_index is not None: data["forecast_dispersion_index"] = f"{weather.dispersion_index:g}"
    if weather.kbdi is not None: data["forecast_kbdi"] = f"{weather.kbdi:g}"
    def write_cell_safe(cell_ref: str, value: Any) -> None:
        target = ws[cell_ref]
        if isinstance(target, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell_ref in merged_range:
                    ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value; return
        target.value = value
    for key, cell in CELL_MAP.items():
        if key in data and data[key] not in (None, ""):
            write_cell_safe(cell, data[key])
    if "AI Rule Check" in wb.sheetnames: del wb["AI Rule Check"]
    chk = wb.create_sheet("AI Rule Check")
    chk.append(["Status", "Item", "Note"])
    for row in build_rule_check(weather): chk.append(list(row))
    chk.append([]); chk.append(["Important", "Human review required", "Draft only. Qualified burn manager must verify field conditions, permits, smoke management, and go/no-go decision."])
    if "NWS Forecast Source" in wb.sheetnames:
        del wb["NWS Forecast Source"]
    source_ws = wb.create_sheet("NWS Forecast Source")
    source_ws.append(["Field", "Value"])
    source_rows = [
        ("Forecast type", "NWS County Fire Weather Planning Forecast (FWF)"),
        ("County", weather.forecast_county),
        ("Forecast period", weather.forecast_period),
        ("NWS office", weather.forecast_office),
        ("Issued", weather.forecast_issued),
        ("NWS product ID", weather.forecast_product_id),
        ("Chance of precipitation", f"{weather.chance_precip_pct:g}%" if weather.chance_precip_pct is not None else ""),
        ("Precipitation type", weather.precip_type),
        ("Precipitation amount", weather.precip_amount),
        ("Stability class", weather.stability_class),
        ("Max LVORI", weather.max_lvori if weather.max_lvori is not None else ""),
        ("Dispersion category", weather.dispersion_category),
        ("Remarks", weather.remarks),
        ("Important", "Forecast values are planning information only and do not replace the desired prescription or observed day-of-burn weather."),
    ]
    for row in source_rows:
        source_ws.append(list(row))
    source_ws.column_dimensions["A"].width = 28
    source_ws.column_dimensions["B"].width = 85
    output_path = Path(output_path); output_path.parent.mkdir(parents=True, exist_ok=True); wb.save(output_path); return output_path



def _first_number(value: str) -> float | None:
    match = re.search(r"-?\d+(?:\.\d+)?", value or "")
    return float(match.group(0)) if match else None


def _parse_wind(value: str) -> Tuple[str, float | None]:
    raw = (value or "").strip()
    if not raw:
        return "", None
    if "lgt/var" in raw.lower() or "light" in raw.lower():
        return "Variable", _first_number(raw)
    direction_match = re.match(r"([A-Za-z]{1,3})", raw)
    direction = direction_match.group(1).upper() if direction_match else ""
    return direction, _first_number(raw)


def parse_fwf_county_product(product_text: str, county: str) -> Dict[str, Any]:
    """Parse a county block from an NWS FWF text product.

    FWF is a fixed-width text product. The parser discovers the period-column
    positions from the heading rather than assuming exact widths.
    """
    lines = product_text.replace("\r", "").splitlines()
    county_re = re.compile(rf"^\s*{re.escape(county)}-\s*$", re.IGNORECASE)
    start = next((i for i, line in enumerate(lines) if county_re.match(line)), None)
    if start is None:
        available = []
        for i, line in enumerate(lines[:-1]):
            if re.match(r"^\s*[A-Z]{2}Z\d{3}(?:>\d{3})?-", line) and lines[i + 1].strip().endswith("-"):
                available.append(lines[i + 1].strip().rstrip("-"))
        raise ValueError(
            f"{county} County was not found in this office's FWF product. "
            f"Counties found include: {', '.join(available[:12]) or 'none detected'}."
        )
    end = next((i for i in range(start + 1, len(lines)) if lines[i].strip() == "$$"), len(lines))
    block = lines[start:end]

    first_row_idx = next((i for i, line in enumerate(block) if re.match(r"^\s*Cloud cover\s{2,}", line)), None)
    if first_row_idx is None:
        first_row_idx = next((i for i, line in enumerate(block) if re.match(r"^\s*Temp\s{2,}", line)), None)
    if first_row_idx is None:
        raise ValueError(f"The {county} County FWF block did not contain the expected forecast table.")
    header_idx = first_row_idx - 1
    while header_idx >= 0 and not block[header_idx].strip():
        header_idx -= 1
    header = block[header_idx]
    matches = list(re.finditer(r"\S(?:.*?\S)?(?=\s{2,}|$)", header))
    # The table header has no row label, so every detected group is a forecast period.
    periods = [m.group(0).strip() for m in matches]
    starts = [m.start() for m in matches]
    if not periods:
        raise ValueError("Could not determine forecast periods from the FWF table.")

    parsed: Dict[str, Dict[str, str]] = {p: {} for p in periods}
    remarks = ""
    for line in block[header_idx + 1:]:
        stripped = line.strip()
        if not stripped:
            continue
        if stripped.startswith("Remarks..."):
            remarks = stripped.split("Remarks...", 1)[1].strip()
            continue
        if stripped.startswith("."):
            break
        label = line[:starts[0]].strip()
        if not label:
            continue
        for idx, period in enumerate(periods):
            stop = starts[idx + 1] if idx + 1 < len(starts) else len(line)
            parsed[period][label] = line[starts[idx]:stop].strip()
    for period in periods:
        parsed[period]["Remarks"] = remarks
    return {"county": county, "periods": periods, "data": parsed}


def fetch_county_fwf(county: str, office: str) -> Dict[str, Any]:
    """Retrieve the newest FWF from the NWS API and parse the selected county."""
    office = office.upper().strip()
    headers = {
        "User-Agent": "BurnPlanAI/0.3 (prescribed fire planning application)",
        "Accept": "application/geo+json, application/json",
    }
    index_url = f"https://api.weather.gov/products/types/FWF/locations/{office}"
    index_resp = requests.get(index_url, headers=headers, timeout=25)
    index_resp.raise_for_status()
    graph = index_resp.json().get("@graph", [])
    if not graph:
        raise ValueError(f"No current FWF products were returned for NWS office {office}.")
    newest = graph[0]
    product_id = newest.get("id", "").rstrip("/").split("/")[-1]
    if not product_id:
        raise ValueError(f"The NWS FWF listing for {office} did not include a product ID.")
    product_url = f"https://api.weather.gov/products/{product_id}"
    product_resp = requests.get(product_url, headers=headers, timeout=25)
    product_resp.raise_for_status()

    # NWS product detail responses expose productText at the top level.
    # Some API serializers may also place fields under properties, so support both.
    payload = product_resp.json()
    props = payload.get("properties") or {}
    product_text = (
        payload.get("productText")
        or props.get("productText")
        or payload.get("text")
        or props.get("text")
        or ""
    )
    if not isinstance(product_text, str):
        product_text = str(product_text or "")
    product_text = product_text.strip()
    if not product_text:
        top_level_fields = ", ".join(sorted(payload.keys()))
        raise ValueError(
            f"The newest {office} FWF response did not include readable product text. "
            f"Returned fields: {top_level_fields or 'none'}."
        )

    result = parse_fwf_county_product(product_text, county)
    result.update({
        "office": office,
        "issued": (
            payload.get("issuanceTime")
            or props.get("issuanceTime")
            or newest.get("issuanceTime", "")
        ),
        "product_id": product_id,
        "source_url": product_url,
    })
    return result


def weather_from_fwf_period(result: Dict[str, Any], period: str) -> Dict[str, Any]:
    """Convert one parsed FWF period into app-ready forecast fields."""
    fields = result.get("data", {}).get(period)
    if not fields:
        raise ValueError(f"Forecast period {period!r} is not available.")
    # PM is usually the most relevant daytime 20-ft wind. Fall back to AM.
    surface_raw = fields.get("20ft wnd mph (PM)") or fields.get("20ft wnd mph (AM)") or ""
    surface_dir, surface_speed = _parse_wind(surface_raw)
    transport_dir, transport_speed = _parse_wind(fields.get("Transport wnd (mph)", ""))
    dispersion_raw = fields.get("Dispersion", "")
    dispersion_value = _first_number(dispersion_raw)
    dispersion_category = re.sub(r"^\s*-?\d+(?:\.\d+)?\s*", "", dispersion_raw).strip()
    return {
        "surface_wind_mph": surface_speed,
        "surface_wind_dir": surface_dir,
        "min_rh": _first_number(fields.get("RH %", "")),
        "max_temp_f": _first_number(fields.get("Temp", "")),
        "transport_wind_mph": transport_speed,
        "transport_wind_dir": transport_dir,
        "mixing_height_ft": _first_number(fields.get("Mixing hgt (ft-AGL)", "")),
        "dispersion_index": dispersion_value,
        "forecast_period": period,
        "forecast_county": result.get("county", ""),
        "forecast_office": result.get("office", ""),
        "forecast_issued": result.get("issued", ""),
        "forecast_product_id": result.get("product_id", ""),
        "chance_precip_pct": _first_number(fields.get("Chance precip (%)", "")),
        "precip_type": fields.get("Precip type", ""),
        "precip_amount": fields.get("Precip amount", ""),
        "stability_class": fields.get("Stability class", ""),
        "max_lvori": _first_number(fields.get("Max LVORI", "")),
        "dispersion_category": dispersion_category,
        "remarks": fields.get("Remarks", ""),
        "raw_fields": fields,
    }

def nws_point_metadata(latitude: float, longitude: float) -> Dict[str, Any]:
    r = requests.get(f"https://api.weather.gov/points/{latitude},{longitude}", headers={"User-Agent": "BurnPlanAI/0.1"}, timeout=20)
    r.raise_for_status(); return r.json().get("properties", {})

def export_pdf(inputs: BurnInputs, weather: WeatherInputs, output_path: str | Path, use_ai: bool = False) -> Path:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    output_path = Path(output_path); output_path.parent.mkdir(parents=True, exist_ok=True)
    draft = basic_ai_draft(inputs, weather)
    if use_ai: draft = optional_openai_polish(draft, inputs, weather)
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="CoverTitle", parent=styles["Title"], fontSize=24, leading=30, alignment=TA_CENTER, spaceAfter=18))
    styles.add(ParagraphStyle(name="SectionHeader", parent=styles["Heading2"], fontSize=14, leading=18, textColor=colors.HexColor("#1f3b2d"), spaceBefore=12, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=9, leading=12))
    def clean(v: Any) -> str:
        return "" if v is None else str(v).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    def para(text: Any, style="BodyText"): return Paragraph(clean(text), styles[style])
    def section(title: str): return para(title, "SectionHeader")
    def table(rows, widths=None):
        t = Table([[para(c, "Small") for c in row] for row in rows], colWidths=widths)
        t.setStyle(TableStyle([("GRID", (0,0), (-1,-1), 0.25, colors.lightgrey), ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#e8efe9")), ("VALIGN", (0,0), (-1,-1), "TOP"), ("LEFTPADDING", (0,0), (-1,-1), 6), ("RIGHTPADDING", (0,0), (-1,-1), 6), ("TOPPADDING", (0,0), (-1,-1), 5), ("BOTTOMPADDING", (0,0), (-1,-1), 5)]))
        return t
    def footer(canvas, doc):
        canvas.saveState(); canvas.setFont("Helvetica", 8); canvas.drawString(0.55*inch, 0.35*inch, f"BurnPlan AI - {inputs.tract_name or 'Draft'}"); canvas.drawRightString(7.95*inch, 0.35*inch, f"Page {doc.page}"); canvas.restoreState()
    doc = SimpleDocTemplate(str(output_path), pagesize=letter, rightMargin=0.55*inch, leftMargin=0.55*inch, topMargin=0.6*inch, bottomMargin=0.6*inch)
    story: List[Any] = []
    story += [Spacer(1, 0.8*inch), para("Prescribed Burn Plan", "CoverTitle"), para(inputs.tract_name or "Draft Burn Unit", "Title"), Spacer(1, 0.2*inch)]
    story.append(table([["County", inputs.county], ["State", inputs.state], ["Acres", f"{inputs.burn_acres:g}" if inputs.burn_acres else ""], ["Burn Type", inputs.burn_type], ["Burn Manager", inputs.burn_mgr_name], ["Date Prepared", datetime.now().strftime("%m/%d/%Y")]], widths=[2.1*inch, 4.7*inch]))
    story += [Spacer(1, 0.25*inch), para("Draft only. Final review, permitting, site verification, weather verification, smoke screening, and go/no-go decisions remain the responsibility of the qualified burn manager.", "Small"), PageBreak()]
    story.append(section("1. Project Information")); story.append(table([["Tract / Burn Unit", inputs.tract_name], ["Location", inputs.burn_address], ["County / State", f"{inputs.county}, {inputs.state}"], ["Latitude / Longitude", "" if inputs.latitude is None or inputs.longitude is None else f"{inputs.latitude:.6f}, {inputs.longitude:.6f}"], ["Burn Acres", f"{inputs.burn_acres:g}" if inputs.burn_acres else ""], ["Burn Type", inputs.burn_type]], widths=[2*inch, 4.8*inch]))
    story.append(section("2. Ownership & Contacts")); story.append(table([["Burn Manager", inputs.burn_mgr_name], ["Certification #", inputs.burn_mgr_cert], ["Phone", inputs.burn_mgr_phone], ["Executor / Landowner Address", inputs.executers_mailing_address], ["Neighbors / Notifications", inputs.neighbors]], widths=[2*inch, 4.8*inch]))
    t = get_prescription_template(inputs.burn_type)
    story.append(section("3. Prescription Engine Recommendation")); story.append(table([["Item", "Recommendation"], ["Season", t.get("season", "")], ["Temperature", inputs.desired_temperature], ["RH", inputs.desired_humidity], ["Surface Wind", inputs.desired_surface_wind], ["Transport Wind", inputs.desired_transport_wind], ["Mixing Height", inputs.desired_mixing_height], ["Dispersion", inputs.desired_dispersion_index], ["Fine Fuel Moisture", inputs.desired_fine_fuel_moisture], ["KBDI", inputs.desired_kbdi], ["Flame Length", inputs.flame_length or t.get("flame_length", "")], ["Nighttime Viable", t.get("nighttime_viable", "")], ["Notes", t.get("notes", "")]], widths=[2*inch, 4.8*inch]))
    story.append(section("4. Objectives")); story.append(para(draft.get("objectives", inputs.objectives))); story.append(Spacer(1, 6)); story.append(para(f"Special Features to Protect: {draft.get('special_features', inputs.special_features)}"))
    story.append(section("5. Burn Unit Description")); story.append(table([["Overstory", inputs.overstory_type], ["Understory", inputs.understory_type], ["Fuel Type / Load", inputs.fuel_type_amount], ["Topography", inputs.topography], ["Firebreaks", inputs.roads_access], ["Water Sources", inputs.water_sources]], widths=[2*inch, 4.8*inch]))
    story.append(section("6. Weather")); story.append(para("The desired prescription, NWS county forecast, and observed day-of-burn conditions are separate records. Imported forecast values do not alter the approved prescription.", "Small")); story.append(Spacer(1, 5)); story.append(table([["Item", "Desired Prescription", "NWS County Forecast", "Observed Day-of-Burn"], ["Surface Wind", inputs.desired_surface_wind, f"{weather.surface_wind_mph or ''} mph {weather.surface_wind_dir}".strip(), inputs.observed_surface_wind], ["RH", inputs.desired_humidity, f"{weather.min_rh or ''}%" if weather.min_rh is not None else "", inputs.observed_humidity], ["Temperature", inputs.desired_temperature, f"{weather.max_temp_f or ''} F" if weather.max_temp_f is not None else "", inputs.observed_temperature], ["Transport Wind", inputs.desired_transport_wind, f"{weather.transport_wind_mph or ''} mph {weather.transport_wind_dir}".strip(), inputs.observed_transport_wind], ["Mixing Height", inputs.desired_mixing_height, f"{weather.mixing_height_ft or ''} ft" if weather.mixing_height_ft is not None else "", inputs.observed_mixing_height], ["Dispersion", inputs.desired_dispersion_index, f"{weather.dispersion_index or ''} {weather.dispersion_category}".strip() if weather.dispersion_index is not None else "", inputs.observed_dispersion_index], ["KBDI", inputs.desired_kbdi, f"{weather.kbdi or ''}" if weather.kbdi is not None else "", inputs.observed_kbdi]], widths=[1.25*inch, 1.75*inch, 1.85*inch, 1.95*inch])); story.append(Spacer(1, 5)); story.append(table([["NWS County Forecast Source", "Value"], ["County / Period", f"{weather.forecast_county} - {weather.forecast_period}".strip(" -")], ["Forecast Office", weather.forecast_office], ["Issued", weather.forecast_issued], ["Product ID", weather.forecast_product_id], ["Chance Precipitation", f"{weather.chance_precip_pct:g}%" if weather.chance_precip_pct is not None else ""], ["Precipitation", " ".join(x for x in [weather.precip_type, weather.precip_amount] if x)], ["Stability / LVORI", " / ".join(x for x in [weather.stability_class, f"LVORI {weather.max_lvori:g}" if weather.max_lvori is not None else ""] if x)], ["Remarks", weather.remarks]], widths=[2*inch, 4.8*inch]))
    story.append(section("7. Smoke & Precautions")); story.append(table([["Smoke Sensitive Areas", inputs.smoke_sensitive_areas], ["Nighttime Smoke Screening", inputs.nighttime_smoke_screening], ["Special Precautions", inputs.special_precautions], ["Smoke Plan", draft.get("smoke_precautions", inputs.smoke_precautions)]], widths=[2*inch, 4.8*inch]))
    story.append(section("8. Personnel & Equipment")); story.append(para(draft.get("manpower_equipment", inputs.manpower_equipment)))
    story.append(section("9. Ignition, Holding & Contingency")); story.append(para(draft.get("ignition_techniques", inputs.ignition_techniques))); story.append(Spacer(1, 6)); story.append(para(f"Breach Potential / Escape Risk: {draft.get('breach_potential', inputs.breach_potential)}")); story.append(Spacer(1, 6)); story.append(para(draft.get("emergency_resources", inputs.emergency_resources)))
    story.append(section("10. Final Burn Record")); story.append(table([["Permit #", inputs.permit_number], ["Actual Burn Date", inputs.actual_burn_date], ["Start Time", inputs.start_time], ["Completion Time", inputs.completion_time], ["Hours to Complete", inputs.hours_to_complete], ["Observed / Expected Flame Length", inputs.flame_length]], widths=[2*inch, 4.8*inch]))
    story.append(section("11. Plan Approval")); story.append(table([["Prepared By", "Signature", "Date"], [inputs.prepared_by_name, "______________________________", inputs.prepared_by_date], ["Witnessed By", "Signature", "Date"], [inputs.witnessed_by_name, "______________________________", inputs.witnessed_by_date]], widths=[2.2*inch, 2.7*inch, 1.9*inch]))
    story.append(section("Rule Check")); story.append(table([["Status", "Item", "Note"]] + build_rule_check(weather), widths=[1*inch, 1.6*inch, 4.2*inch]))
    doc.build(story, onFirstPage=footer, onLaterPages=footer); return output_path
